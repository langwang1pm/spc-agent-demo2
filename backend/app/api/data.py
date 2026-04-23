from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List
from app.core.database import get_db
from app.models import DataSource, DataSourceType, SystemSourceType
from app.schemas import (
    DataSourceCreate,
    DataSourceResponse,
    ManualDataInput,
    ApiResponse
)
import shutil
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter(prefix="/data", tags=["数据输入"])

# 文件上传目录
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)


@router.get("/template", response_class=FileResponse)
async def download_template():
    """下载数据导入模板"""
    from openpyxl.comments import Comment

    wb = Workbook()
    ws = wb.active
    ws.title = "检验记录表A"

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    label_font = Font(bold=True, size=11)
    label_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7")
    )
    placeholder_font = Font(color="BFBFBF", size=10, italic=True)

    NUM_SAMPLES = 6  # B~G列
    NUM_ITEMS = 6    # 2~7行

    # --- 第1行表头 ---
    ws.cell(row=1, column=1, value=None)  # A1 空白
    ws.cell(row=1, column=1).border = thin_border

    for col_idx in range(2, 2 + NUM_SAMPLES):
        cell = ws.cell(row=1, column=col_idx, value=f"检验样本{col_idx - 1}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # --- A列行标题（检验项）+ 数据区域（浅灰占位文本）---
    for row_idx in range(2, 2 + NUM_ITEMS):
        # A列行标题
        label_cell = ws.cell(row=row_idx, column=1, value=f"检验项{row_idx - 1}")
        label_cell.font = label_font
        label_cell.fill = label_fill
        label_cell.alignment = center_align
        label_cell.border = thin_border

        # B~G列数据区域：浅灰色占位示例值
        for col_idx in range(2, 2 + NUM_SAMPLES):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center_align
            cell.border = thin_border
            # 占位示例值（浅灰色斜体，用户输入时自动替换）
            sample_val = round(10.0 + (row_idx % 3) * 0.3 - col_idx * 0.1 + (row_idx * col_idx % 5) * 0.05, 2)
            cell.value = sample_val
            cell.font = placeholder_font
            cell.number_format = '0.00'

    # --- 批注说明 ---
    ws.cell(row=1, column=1).comment = Comment(
        "行 = 检验项（可向下增行）\n列 = 检验样本（可向右增列）\n数据区域填写数值，缺失值留空即可",
        "SPC Agent"
    )

    # --- 列宽 ---
    ws.column_dimensions['A'].width = 12
    for col_idx in range(2, 2 + NUM_SAMPLES):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 14

    # --- 行高 ---
    ws.row_dimensions[1].height = 24
    for r in range(2, 2 + NUM_ITEMS):
        ws.row_dimensions[r].height = 22

    # --- 保存并返回 ---
    template_path = UPLOAD_DIR / "检验数据模板.xlsx"
    wb.save(template_path)
    return FileResponse(
        path=str(template_path),
        filename="检验数据模板.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.post("/manual", response_model=ApiResponse)
async def create_manual_data(
    data: ManualDataInput,
    db: Session = Depends(get_db)
):
    """创建手动输入数据"""
    db_data = DataSource(
        name=data.name,
        source_type=DataSourceType.MANUAL,
        data_values=data.data_values
    )
    db.add(db_data)
    db.commit()
    db.refresh(db_data)
    
    return ApiResponse(
        success=True,
        message="数据添加成功",
        data={"id": db_data.id, "name": db_data.name}
    )


@router.post("/file", response_model=ApiResponse)
async def upload_file_data(
    name: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """上传文件数据（Excel/CSV）——仅保存文件，不解析数据值；SPC计算时实时解析"""
    # 检查文件类型
    if not (file.filename.endswith('.xlsx') or 
            file.filename.endswith('.xls') or 
            file.filename.endswith('.csv')):
        raise HTTPException(status_code=400, detail="只支持Excel(.xlsx/.xls)或CSV文件")
    
    # 保存文件（仅存储文件，不解析数据值）
    file_path = UPLOAD_DIR / f"{file.filename}"
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # 创建数据源记录（只存文件路径，data_values 留空）
    db_data = DataSource(
        name=name,
        source_type=DataSourceType.FILE,
        file_name=file.filename,
        file_path=str(file_path)
    )
    db.add(db_data)
    db.commit()
    db.refresh(db_data)
    
    return ApiResponse(
        success=True,
        message="文件上传成功",
        data={"id": db_data.id, "name": db_data.name, "file_name": file.filename}
    )


@router.post("/system", response_model=ApiResponse)
async def create_system_data(
    data: DataSourceCreate,
    db: Session = Depends(get_db)
):
    """创建系统对接数据源"""
    if data.source_type != DataSourceType.SYSTEM:
        raise HTTPException(status_code=400, detail="数据源类型必须为system")
    
    db_data = DataSource(
        name=data.name,
        source_type=DataSourceType.SYSTEM,
        system_type=data.system_type,
        connection_config=data.connection_config,
        query_config=data.query_config
    )
    db.add(db_data)
    db.commit()
    db.refresh(db_data)
    
    return ApiResponse(
        success=True,
        message="系统对接配置创建成功",
        data={"id": db_data.id, "name": db_data.name}
    )


@router.get("/list", response_model=ApiResponse)
async def list_data_sources(
    skip: int = 0,
    limit: int = 20,
    db: Session = Depends(get_db)
):
    """获取数据源列表"""
    total = db.query(DataSource).count()
    data_sources = db.query(DataSource).order_by(
        DataSource.created_at.desc()
    ).offset(skip).limit(limit).all()
    
    return ApiResponse(
        success=True,
        data={
            "total": total,
            "items": [
                {
                    "id": ds.id,
                    "name": ds.name,
                    "source_type": ds.source_type.value,
                    "created_at": ds.created_at.isoformat() if ds.created_at else None
                }
                for ds in data_sources
            ]
        }
    )


@router.get("/{data_id}", response_model=ApiResponse)
async def get_data_source(
    data_id: int,
    db: Session = Depends(get_db)
):
    """获取单个数据源详情"""
    data_source = db.query(DataSource).filter(DataSource.id == data_id).first()
    if not data_source:
        raise HTTPException(status_code=404, detail="数据源不存在")
    
    return ApiResponse(
        success=True,
        data={
            "id": data_source.id,
            "name": data_source.name,
            "source_type": data_source.source_type.value,
            "data_values": data_source.data_values,
            "file_name": data_source.file_name,
            "file_path": data_source.file_path,
            "system_type": data_source.system_type.value if data_source.system_type else None,
            "connection_config": data_source.connection_config,
            "query_config": data_source.query_config,
            "created_at": data_source.created_at.isoformat() if data_source.created_at else None,
            "updated_at": data_source.updated_at.isoformat() if data_source.updated_at else None
        }
    )


@router.delete("/{data_id}", response_model=ApiResponse)
async def delete_data_source(
    data_id: int,
    db: Session = Depends(get_db)
):
    """删除数据源"""
    data_source = db.query(DataSource).filter(DataSource.id == data_id).first()
    if not data_source:
        raise HTTPException(status_code=404, detail="数据源不存在")
    
    # 删除关联文件
    if data_source.file_path and os.path.exists(data_source.file_path):
        os.remove(data_source.file_path)
    
    db.delete(data_source)
    db.commit()
    
    return ApiResponse(success=True, message="数据源删除成功")