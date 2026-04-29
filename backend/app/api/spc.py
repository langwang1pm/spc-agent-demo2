"""
SPC分析和导出相关API
"""
from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, List
from pathlib import Path
import io, csv, json
from openpyxl import load_workbook
# pandas removed — no longer used
# import pandas as pd

from app.core.database import get_db
from app.models import DataSource, DataSourceType, AnalysisConfig, AIAnalysisRecord, SystemSettings
from app.schemas import AnalysisConfigCreate, ApiResponse
from app.services.spc import calculate_spc
from app.services.ai_agent import get_ai_analysis
from app.services.export import export_service
from app.services.system_query import get_system_data_values, SystemQueryError


def _convert_to_2d(data_values: List[float], subgroup_size: int) -> List[List[float]]:
    """
    将一维数组按子组大小转换为二维数组。
    不能整除时，最后一组保留实际数据量（不补NaN，不丢弃）。
    例如：36个数据，subgroup_size=5 → 8组(5+5+5+5+5+5+5+1)
    """
    result = []
    for i in range(0, len(data_values), subgroup_size):
        group = data_values[i:i + subgroup_size]
        result.append(group)
    return result


def _get_data_values_from_source(
    data_source: DataSource,
    subgroup_size: int = 5
) -> List[List[float]]:
    """
    从数据源获取数据值（统一处理三种数据源类型）
    
    Args:
        data_source: 数据源对象
        subgroup_size: 子组大小（用于一维转二维）
        
    Returns:
        二维数据数组
        
    Raises:
        HTTPException: 数据获取失败时抛出
    """
    if data_source.source_type == DataSourceType.FILE:
        if not data_source.file_path:
            raise HTTPException(status_code=400, detail="文件数据源未找到文件路径")
        try:
            return _parse_file_values(data_source.file_path)
        except FileNotFoundError as e:
            raise HTTPException(status_code=400, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")
    
    elif data_source.source_type == DataSourceType.SYSTEM:
        # 系统对接数据源
        if not data_source.system_type:
            raise HTTPException(status_code=400, detail="系统对接数据源未指定系统类型")
        if not data_source.connection_config:
            raise HTTPException(status_code=400, detail="系统对接数据源未配置连接信息")
        if not data_source.query_config:
            raise HTTPException(status_code=400, detail="系统对接数据源未配置查询语句")
        
        try:
            # 从外部系统查询数据
            raw_values = get_system_data_values(
                system_type=data_source.system_type.value,
                connection_config=data_source.connection_config,
                query_config=data_source.query_config
            )
            # 将一维数组转换为二维数组（按子组大小分组）
            return _convert_to_2d(raw_values, subgroup_size)
        except SystemQueryError as e:
            raise HTTPException(status_code=400, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"系统数据查询失败: {str(e)}")
    
    else:
        # MANUAL：直接使用存储在 data_values 中的数据
        if not data_source.data_values:
            raise HTTPException(status_code=400, detail="数据源无有效数据")
        # PostgreSQL JSONB 返回的可能是字符串，需解析
        raw_data = data_source.data_values
        if isinstance(raw_data, str):
            raw_data = json.loads(raw_data)
        # 检测数据维度：
        if isinstance(raw_data[0], (int, float)):
            return _convert_to_2d(raw_data, subgroup_size)
        else:
            return raw_data  # 已是二维数组，直接使用


def _parse_file_values(file_path: str) -> List[List[float]]:
    """
    根据文件路径解析 Excel/CSV，返回 List[List[float]]。
    模板格式：第1行表头，A列为行标题，数据区B2开始。
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    if path.suffix.lower() == '.csv':
        rows = []
        with open(path, newline='', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row in reader:
                # 跳过表头行（首行）
                if reader.line_num == 1:
                    continue
                row_vals = []
                for v in row:
                    try:
                        fv = float(v)
                        if not (fv != fv):  # 排除 NaN
                            row_vals.append(fv)
                    except (ValueError, TypeError):
                        pass
                if row_vals:
                    rows.append(row_vals)
    else:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_vals = []
            for v in row:
                if v is None:
                    continue
                try:
                    fv = float(v)
                    if not (fv != fv):  # 排除 NaN
                        row_vals.append(fv)
                except (ValueError, TypeError):
                    pass
            if row_vals:
                rows.append(row_vals)
        wb.close()
    return rows

router = APIRouter(prefix="/spc", tags=["SPC分析"])


@router.get("/calculate", response_model=ApiResponse)
async def calculate_spc_chart(
    data_source_id: int,
    chart_type: str = "xbar_r",
    subgroup_size: int = 5,
    confidence_level: str = "99",
    show_rules: bool = True,
    show_prediction: bool = False,
    db: Session = Depends(get_db)
):
    """
    执行SPC计算并生成图表数据
    
    Args:
        data_source_id: 数据源ID
        chart_type: 图表类型
        subgroup_size: 子组大小
        confidence_level: 置信水平
        show_rules: 是否显示判异规则
        show_prediction: 是否显示预测区间
    """
    # 获取数据源
    data_source = db.query(DataSource).filter(DataSource.id == data_source_id).first()
    if not data_source:
        raise HTTPException(status_code=404, detail="数据源不存在")

    # 根据数据源类型获取数据值
    data_values = _get_data_values_from_source(data_source, subgroup_size)
    if not data_values:
        raise HTTPException(status_code=400, detail="数据源中未找到有效数值数据")

    # 执行SPC计算
    result = calculate_spc(
        data=data_values,
        chart_type=chart_type,
        subgroup_size=subgroup_size,
        confidence_level=confidence_level,
        show_rules=show_rules,
        show_prediction=show_prediction
    )
    
    return ApiResponse(
        success=True,
        message="SPC计算完成",
        data={
            "chart_type": result["chart_type"],
            "chart_data": result["chart_data"],
            "data_values": result["data_values"],
            "statistics": result["statistics"],
            "control_limits": result["control_limits"],
            "anomalies": result["anomalies"],
            "rules_violations": result["rules_violations"]
        }
    )


@router.post("/analyze", response_model=ApiResponse)
async def analyze_with_ai(
    data_source_id: int,
    analysis_config_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """调用AI进行分析"""
    data_source = db.query(DataSource).filter(DataSource.id == data_source_id).first()
    if not data_source:
        raise HTTPException(status_code=404, detail="数据源不存在")
    
    analysis_config = None
    spc_result = None
    data_values = None

    if analysis_config_id:
        analysis_config = db.query(AnalysisConfig).filter(
            AnalysisConfig.id == analysis_config_id
        ).first()

        # 获取数据值
        try:
            data_values = _get_data_values_from_source(data_source, analysis_config.subgroup_size)
        except HTTPException:
            data_values = None

        if data_values and analysis_config:
            spc_result = calculate_spc(
                data=data_values,
                chart_type=analysis_config.chart_type.value,
                subgroup_size=analysis_config.subgroup_size,
                confidence_level=analysis_config.confidence_level.value,
                show_rules=analysis_config.show_rules,
                show_prediction=analysis_config.show_prediction
            )
    
    # 调用AI分析
    analysis_result = await get_ai_analysis(data_source, analysis_config, spc_result)
    
    # 保存分析记录
    record = AIAnalysisRecord(
        data_source_id=data_source_id,
        analysis_config_id=analysis_config_id,
        analysis_result=analysis_result,
        raw_response={"spc_result": spc_result} if spc_result else None
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    
    return ApiResponse(
        success=True,
        message="AI分析完成",
        data={
            "id": record.id,
            "analysis_result": record.analysis_result,
            "created_at": record.created_at.isoformat()
        }
    )


@router.get("/export/chart", response_class=StreamingResponse)
async def export_spc_chart(
    data_source_id: int,
    chart_type: str = "xbar_r",
    subgroup_size: int = 5,
    confidence_level: str = "99",
    db: Session = Depends(get_db)
):
    """导出SPC图表数据为Excel"""
    data_source = db.query(DataSource).filter(DataSource.id == data_source_id).first()
    if not data_source:
        raise HTTPException(status_code=404, detail="数据源不存在")

    # 获取数据值
    data_values = _get_data_values_from_source(data_source, subgroup_size)
    if not data_values:
        raise HTTPException(status_code=400, detail="数据源中未找到有效数值数据")

    # 计算SPC
    result = calculate_spc(
        data=data_values,
        chart_type=chart_type,
        subgroup_size=subgroup_size,
        confidence_level=confidence_level
    )
    
    # 导出Excel
    excel_data = export_service.export_chart_data(result)
    
    filename = f"SPC_图表数据_{data_source.name}_{chart_type}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )


@router.get("/export/raw", response_class=StreamingResponse)
async def export_raw_data(
    data_source_id: int,
    db: Session = Depends(get_db)
):
    """导出原始数据为Excel"""
    data_source = db.query(DataSource).filter(DataSource.id == data_source_id).first()
    if not data_source:
        raise HTTPException(status_code=404, detail="数据源不存在")

    # 获取数据值
    data_values = _get_data_values_from_source(data_source, subgroup_size=5)
    if not data_values:
        raise HTTPException(status_code=400, detail="数据源中未找到有效数值数据")

    # 导出Excel
    excel_data = export_service.export_raw_data(
        data=data_values,
        data_title=data_source.name
    )
    
    filename = f"SPC_原始数据_{data_source.name}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )


@router.get("/export/report", response_class=StreamingResponse)
async def export_analysis_report(
    data_source_id: int,
    format: str = "docx",
    db: Session = Depends(get_db)
):
    """
    导出AI分析报告
    
    Args:
        data_source_id: 数据源ID
        format: 报告格式 (docx/html/pdf)
    """
    data_source = db.query(DataSource).filter(DataSource.id == data_source_id).first()
    if not data_source:
        raise HTTPException(status_code=404, detail="数据源不存在")
    
    # 获取最新的分析记录
    record = db.query(AIAnalysisRecord).filter(
        AIAnalysisRecord.data_source_id == data_source_id
    ).order_by(AIAnalysisRecord.created_at.desc()).first()
    
    if not record or not record.analysis_result:
        raise HTTPException(status_code=404, detail="未找到分析记录，请先执行AI分析")
    
    data_info = {
        "data_title": data_source.name,
        "source_type": data_source.source_type.value
    }
    
    # 根据格式导出
    if format == "docx":
        content_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        filename = f"SPC_分析报告_{data_source.name}.docx"
    elif format == "html":
        content_type = "text/html"
        filename = f"SPC_分析报告_{data_source.name}.html"
    elif format == "pdf":
        content_type = "application/pdf"
        filename = f"SPC_分析报告_{data_source.name}.pdf"
    else:
        raise HTTPException(status_code=400, detail=f"不支持的格式: {format}")
    
    report_data = export_service.export_analysis_report(
        analysis_content=record.analysis_result,
        data_info=data_info,
        report_format=format
    )
    
    return StreamingResponse(
        io.BytesIO(report_data),
        media_type=content_type,
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )